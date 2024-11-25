"FP&A Report Generator GUI"


import logging
import pandas as pd
from pathlib import Path
from tkinter import Tk, Button, Label, Text, Frame, Toplevel, OptionMenu, StringVar, Entry
from tkinter.filedialog import askdirectory, askopenfilename
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
import re  
import webbrowser
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import csv

# Configure logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    filename='process.log',
    filemode='w'  # Overwrite log file on each run
)


# Helper function to extract date from folder names
def extract_date_from_path(path):
    match = re.search(r'\b\d{2}-\d{2}-\d{2}\b', str(path))
    if match:
        try:
            return pd.to_datetime(match.group(), format="%d-%m-%y").date()
        except ValueError as e:
            logging.error(f"Error parsing date from path {path}: {e}")
            return None
    return None


# Function to load data from folder
def load_data(folder):
    logging.info(f"Loading data from folder: {folder}")
    total_transactions = pd.DataFrame(columns=["Client", "Country", "Currency", "Transaction", "Date"])
    folder_path = Path(folder)

    for file in folder_path.rglob("*"):
        if not file.is_file() or file.suffix not in [".xlsx", ".csv"]:
            logging.debug(f"Skipping non-data file: {file}")
            continue

        transaction_date = extract_date_from_path(file)
        if transaction_date is None:
            logging.warning(f"Skipping file {file} - No valid date found in path.")
            continue

        try:
            if file.suffix == ".xlsx":
                logging.debug(f"Loading Excel file: {file}")
                data = pd.read_excel(file, usecols=["Client", "Country", "Currency", "Transaction"])
            elif file.suffix == ".csv":
                logging.debug(f"Loading CSV file: {file}")
                with open(file, 'r', newline='', encoding='utf-8') as f:
                    sniffer = csv.Sniffer()
                    sample = f.read(2048)
                    f.seek(0)
                    delimiter = sniffer.sniff(sample).delimiter
                data = pd.read_csv(file, delimiter=delimiter, on_bad_lines='skip')
            
            # Normalize column names and add the transaction date
            data.columns = data.columns.str.strip().str.title()
            data["Date"] = transaction_date
            
            # Ensure columns are consistent with total_transactions
            data = data[["Client", "Country", "Currency", "Transaction", "Date"]]
            if not data.empty:  
                total_transactions = pd.concat([total_transactions, data], ignore_index=True)
            logging.info(f"Successfully loaded file: {file}")
        except Exception as e:
            logging.error(f"Error reading file {file}: {e}")

    total_transactions["Transaction"] = total_transactions["Transaction"].round(2) 
    logging.info("Data loading completed.")
    return total_transactions


# Function to calculate transactions in USD
def calculate_usd_transactions(data, rates):
    logging.info("Calculating transactions in USD.")
    try:
        data_with_rates = data.merge(rates, on="Currency", how="left")
        missing_rates = data_with_rates[data_with_rates["Rate"].isnull()]["Currency"].unique()
        if len(missing_rates) > 0:
            logging.warning(f"Missing conversion rates for currencies: {missing_rates}")
        data_with_rates["Transaction USD"] = data_with_rates["Transaction"] * data_with_rates["Rate"]
        data_with_rates["Transaction USD"] = data_with_rates["Transaction USD"].fillna(0).round(2)
        logging.info("USD transaction calculation completed.")
        return data_with_rates
    except Exception as e:
        logging.error(f"Error calculating USD transactions: {e}")
        return data


# Function to load conversion rates
def load_conversion_rates(file_path):
    logging.info(f"Loading conversion rates from: {file_path}")
    try:
        rates = pd.read_excel(file_path, header=6)
        rates.columns = rates.columns.str.lower()
        if "code" not in rates.columns or "rate" not in rates.columns:
            raise KeyError("The rates file does not contain the expected columns ('Code', 'Rate').")
        rates = rates.rename(columns={"code": "Currency", "rate": "Rate"})
        rates = rates[["Currency", "Rate"]].dropna()
        rates["Rate"] = pd.to_numeric(rates["Rate"], errors="coerce").dropna()
        rates["Rate"] = rates["Rate"].round(2)
        logging.info("Conversion rates loaded successfully.")
        return rates
    except Exception as e:
        logging.error(f"Error loading conversion rates: {e}")
        return pd.DataFrame()


def create_pivot_tables(data, wb):
    # Create a table for "Revenue by Market Section"
    pivot_market_section = data.copy()
    pivot_market_section["Market Section"] = pivot_market_section["Client"].str[:1]  # Obtaining the first digit for client name
    pivot_market_section = pivot_market_section.groupby(["Market Section", "Date"], as_index=False)["Transaction USD"].sum()

    # Create a table for "Revenue by Geography"
    pivot_geography = data.groupby("Country", as_index=False)["Transaction USD"].sum()

    # Add tables to the wb
    pivot_ws1 = wb.create_sheet(title="Revenue by Market Section")
    pivot_ws2 = wb.create_sheet(title="Revenue by Country")

    # Insert headers
    pivot_ws1.append(["Market Section", "Date", "Total Revenue (USD)"])
    for r in dataframe_to_rows(pivot_market_section, index=False, header=False):
        pivot_ws1.append(r)

    # Format "Revenue by Market Section"
    table_ref_market = f"A2:C{len(pivot_market_section) + 2}"
    table_market = Table(displayName="MarketSectionTable", ref=table_ref_market)
    table_market.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    pivot_ws1.add_table(table_market)

    # Insertar header
    pivot_ws2.append(["Country", "Total Revenue (USD)"])
    for r in dataframe_to_rows(pivot_geography, index=False, header=False):
        pivot_ws2.append(r)

    # Format tables
    table_ref_country = f"A2:B{len(pivot_geography) + 2}"
    table_country = Table(displayName="GeographyTable", ref=table_ref_country)
    table_country.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    pivot_ws2.add_table(table_country)


# Function to generate the report in Excel
def generate_report(original_data, data_with_usd, label, text, button, filter_button=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Add header with formatting
    add_excel_header(ws, "Total Transactions Report", len(original_data.columns))

    # Append data
    for r in dataframe_to_rows(original_data, index=False, header=True):
        ws.append(r)

    # Format table
    table_ref = f"A5:{get_column_letter(len(original_data.columns))}{len(original_data) + 5}"
    table = Table(displayName="TransactionTable", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    ws.add_table(table)

    # Totals in USD sheet
    pivot_ws = wb.create_sheet(title="Totals in USD")
    add_excel_header(pivot_ws, "Total Transactions by Client in USD", 2)

    # Group data and append
    pivot_data = data_with_usd.groupby("Client", as_index=False)["Transaction USD"].sum()
    headers = ["Client", "Sum of Transaction USD"]
    pivot_ws.append(headers)
    for r in dataframe_to_rows(pivot_data, index=False, header=False):
        pivot_ws.append(r)

    # Format table
    table_ref_usd = f"A5:{get_column_letter(len(pivot_data.columns))}{len(pivot_data) + 5}"
    table_usd = Table(displayName="USDTransactionTable", ref=table_ref_usd)
    table_usd.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    pivot_ws.add_table(table_usd)

    # Revenue by Market Section sheet
    pivot_ws_market = wb.create_sheet(title="Revenue by Market Section")
    add_excel_header(pivot_ws_market, "Revenue by Market Section", 3)

    # Generate and append data for Market Section by Date
    market_section_data = data_with_usd.copy()
    market_section_data["Market Section"] = market_section_data["Client"].str[:1] 
    market_section_grouped = market_section_data.groupby(
        ["Market Section", "Date"], as_index=False
    )["Transaction USD"].sum()

    # Add headers and rows
    pivot_ws_market.append(["Market Section", "Date", "Total Revenue (USD)"])
    for r in dataframe_to_rows(market_section_grouped, index=False, header=False):
        pivot_ws_market.append(r)

    # Format table for Market Section
    table_ref_market = f"A5:C{len(market_section_grouped) + 5}"
    table_market = Table(displayName="MarketSectionTable", ref=table_ref_market)
    table_market.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    pivot_ws_market.add_table(table_market)

    # Revenue by Country sheet
    pivot_ws_country = wb.create_sheet(title="Revenue by Country")
    add_excel_header(pivot_ws_country, "Revenue by Country", 2)

    # Generate and append data for Geography
    geography_grouped = data_with_usd.groupby("Country", as_index=False)["Transaction USD"].sum()
    pivot_ws_country.append(["Country", "Total Revenue (USD)"])
    for r in dataframe_to_rows(geography_grouped, index=False, header=False):
        pivot_ws_country.append(r)

    # Format table for Geography
    table_ref_country = f"A5:B{len(geography_grouped) + 5}"
    table_country = Table(displayName="GeographyTable", ref=table_ref_country)
    table_country.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    pivot_ws_country.add_table(table_country)

    # Adjust column widths
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            try:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_length + 2
            except Exception as e:
                logging.error(f"Error adjusting column {col}: {e}")

    # Disable gridlines for all sheets
    for sheet in wb.sheetnames:
        wb[sheet].sheet_view.showGridLines = False

    # Save the workbook
    wb.save("Marsh_McLennan_Revenue_Report.xlsx")
    
    # Confirmation message
    label.config(text="Report saved as 'Marsh_McLennan_Revenue_Report.xlsx'.")
    text.pack_forget()

    # Clear all widgets
    for widget in button.master.winfo_children():
        widget.pack_forget()

    # Restore the title Label
    Label(button.master, text="Marsh McLennan FP&A Report Generator", font=("Arial", 16, "bold")).pack(pady=10)

    # Add final menu buttons
    Label(button.master, text="Report saved as 'Marsh_McLennan_Revenue_Report.xlsx'.", font=("Arial", 12)).pack(pady=10)
    Button(button.master, text="Send Report Via Email", command=send_email_via_mailto).pack(pady=5)
    Button(button.master, text="Search a Client's Name", command=lambda: search_client(original_data)).pack(pady=5)
    Button(button.master, text="Generate another Report", command=lambda: reset_app(label, text, button)).pack(pady=5)
    Button(button.master, text="Exit", command=button.master.quit).pack(pady=5)


# Function to add formatted header to an Excel sheet
def add_excel_header(sheet, title, num_columns):
    # Merge cells for the header
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_columns)
    sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=num_columns)
    sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=num_columns)

    # Set header values
    sheet["A2"] = "MARSH MCLENNAN"
    sheet["A3"] = title.upper()
    sheet["A4"] = "2024"

    # Define font and alignment
    header_font = Font(bold=True, size=12)
    alignment = Alignment(horizontal="center", vertical="center")

    # Apply styles to header cells
    for cell in ["A2", "A3", "A4"]:
        sheet[cell].font = header_font
        sheet[cell].alignment = alignment

# Function to filter data by date
def filter_by_date(data, text, label, button, filter_button, filtered_data_container):
    # Check if the filter window is already open
    if hasattr(filter_by_date, "filter_window") and filter_by_date.filter_window is not None:
        filter_by_date.filter_window.lift()
        return

    # Create a new filter window
    filter_by_date.filter_window = Toplevel()
    filter_by_date.filter_window.title("Filter by Date")

    # Function to apply the filter
    def apply_filter():
        try:
            start_date = pd.to_datetime(start_date_var.get()).date()
            end_date = pd.to_datetime(end_date_var.get()).date()

            # Filter the data by the selected range
            filtered_data = data[(data["Date"] >= start_date) & (data["Date"] <= end_date)]
            filtered_data_container[0] = filtered_data  # Store the filtered data for later use
            
            # Display filtered data in the main text widget
            label.config(text=f"Data filtered from {start_date} to {end_date}.")
            text.delete('1.0', 'end')
            text.insert('1.0', filtered_data.to_string())
            text.pack()

            # Close the filter window after filtering
            filter_by_date.filter_window.destroy()
            filter_by_date.filter_window = None
        except Exception as e:
            label.config(text=f"Error applying filter: {e}")

    # Dropdown for start date
    Label(filter_by_date.filter_window, text="Select Start Date").pack(pady=5)
    start_date_var = StringVar(filter_by_date.filter_window)
    start_date_var.set(str(data["Date"].min()))
    OptionMenu(filter_by_date.filter_window, start_date_var, *sorted(data["Date"].astype(str).unique())).pack(pady=5)

    # Dropdown for end date
    Label(filter_by_date.filter_window, text="Select End Date").pack(pady=5)
    end_date_var = StringVar(filter_by_date.filter_window)
    end_date_var.set(str(data["Date"].max()))
    OptionMenu(filter_by_date.filter_window, end_date_var, *sorted(data["Date"].astype(str).unique())).pack(pady=5)

    # Button to apply the filter
    Button(filter_by_date.filter_window, text="Apply Filter", command=apply_filter).pack(pady=10)

    # Close window properly when user closes it manually
    filter_by_date.filter_window.protocol("WM_DELETE_WINDOW", lambda: setattr(filter_by_date, "filter_window", None))


# Function to select a rates file
def select_rates_file(label, text, data, button, filter_button, filtered_data_container):
    file_path = askopenfilename(
        title="Select the conversion rates file",
        filetypes=[("Excel Files", "*.xlsx")]
    )
    if file_path:
        rates = load_conversion_rates(file_path)
        if not rates.empty:
            text.delete('1.0', 'end')
            text.insert('1.0', rates.to_string(index=False))
            text.pack()

            # Use filtered data if available, otherwise use original data
            data_to_use = filtered_data_container[0] if filtered_data_container[0] is not None else data
            data_with_usd = calculate_usd_transactions(data_to_use, rates)

            label.config(text="Rates loaded successfully. You can now generate the report.")
            button.config(
                text="Generate Report",
                command=lambda: generate_report(data_to_use, data_with_usd, label, text, button, filter_button)
            )

            # Hide the filter button
            if filter_button:
                filter_button.pack_forget()
        else:
            label.config(text="Error loading conversion rates.")
    else:
        label.config(text="No rates file was selected.")


# Function for folder selection workflow
def on_folder_selected(label, text, button):
    folder = select_folder()
    if folder:
        data = load_data(folder)
        if not data.empty:
            label.config(text="Data loaded successfully. You can now filter data by date or select a rates file to generate the report.")
            text.delete('1.0', 'end')
            text.insert('1.0', data.to_string())
            text.pack()

            # Create a container to store the filtered data
            filtered_data_container = [None]

            # Update "Select Rates File" button to include the filtered data container
            button.config(
                text="Select Rates File",
                command=lambda: select_rates_file(label, text, data, button, filter_button, filtered_data_container)
            )

            # Create filter button and place it above the "Select Rates File" button
            filter_button = Button(
                button.master,
                text="Filter by Date",
                command=lambda: filter_by_date(data, text, label, button, filter_button, filtered_data_container)
            )
            filter_button.pack(pady=5, before=button)

        else:
            label.config(text="No data was combined.")
    else:
        label.config(text="No folder was selected.")


#Send Report
def send_email_via_mailto():
    recipient = "recipient@example.com"
    subject = "Marsh McLennan Revenue Report"
    body = "Please find attached the Revenue Report. (Remember to attach the file before sending)"

    # Format the mailto URL
    mailto_url = f"mailto:{recipient}?subject={subject}&body={body}"
    
    # Open the default mail client
    webbrowser.open(mailto_url)


#Selenium Robot to search client's name
def search_client(original_data):
    # Create a new window (Toplevel)
    search_window = Toplevel()
    search_window.title("Search Client")
    search_window.geometry("400x200")

    # Label to show instructions
    Label(search_window, text="Enter Client Name:", font=("Arial", 12)).pack(pady=10)

    # Variable to capture client's name
    client_name_var = StringVar()
    Entry(search_window, textvariable=client_name_var, font=("Arial", 12), width=30).pack(pady=5)

    # Search label
    search_label = Label(search_window, text="", font=("Arial", 12))
    search_label.pack(pady=5)

    # Internal function to perform the search
    def perform_search():
        client_name = client_name_var.get()  # Getting client's name
        if client_name in original_data["Client"].values:
            try:
                logging.info(f"Searching for client: {client_name}")

                # Chrome settingd
                chrome_options = Options()
                chrome_options.add_argument("--start-maximized")
                chrome_options.add_argument("--disable-infobars")
                chrome_options.add_argument("--disable-extensions")

                # Initialize Chrome with webdriver_manager
                service = Service(ChromeDriverManager().install())
                driver = webdriver.Chrome(service=service, options=chrome_options)

                # Perform the search
                driver.get("https://www.google.com")
                search_box = driver.find_element("name", "q")
                search_box.send_keys(client_name)
                search_box.send_keys(Keys.RETURN)
                logging.info("Search completed. Browser will remain open.")

            except Exception as e:
                logging.error(f"Error while performing search: {e}")
                search_label.config(text="An error occurred during the search.", fg="red")

    Button(search_window, text="Search", command=perform_search).pack(pady=10)


#Reset app
def reset_app(label, text, button):
    # Clear all widgets except the title
    for widget in button.master.winfo_children():
        widget.pack_forget()

    # Restore the title Label
    Label(button.master, text="Marsh McLennan FP&A Report Generator", font=("Arial", 16, "bold")).pack(pady=10)

    # Restore the initial widgets
    label.config(text="Select a folder with total transactions for a year:")
    label.pack(pady=5)
    button.config(text="Select Folder", command=lambda: on_folder_selected(label, text, button))
    button.pack(pady=5)
    Button(button.master, text="Exit", command=button.master.quit).pack(pady=5)


# Function to select a folder
def select_folder():
    return askdirectory(title="Select a folder")

# Main GUI function
def create_app():
    logging.info("Starting FP&A Report Generator GUI.")
    root = Tk()
    root.title("FP&A Report Generator")
    frame = Frame(root, padx=10, pady=10)
    frame.pack()

    label_title = Label(frame, text="Marsh McLennan FP&A Report Generator", font=("Arial", 16, "bold"))
    label_title.pack(pady=10)

    text = Text(frame, wrap='word', width=80, height=20)
    label = Label(frame, text="Select a folder with total transactions for a year:")
    label.pack(pady=5)

    button = Button(frame, text="Select Folder", command=lambda: on_folder_selected(label, text, button))
    button.pack(pady=5)

    Button(frame, text="Exit", command=root.destroy).pack(pady=5)
    root.mainloop()
    logging.info("FP&A Report Generator GUI closed.")


if __name__ == "__main__":
    create_app()